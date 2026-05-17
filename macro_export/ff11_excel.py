"""
FF11 マクロ・装備セット・所持アイテム Excel整形スクリプト
使い方:
  python ff11_excel.py [マクロCSV] [装備セットCSV] [所持アイテムCSV]

引数省略時はスクリプトと同じフォルダから最新ファイルを自動検索
"""

import sys
import os
import io
import re
import glob
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv

# ─────────────────────────────────────────
# 色定義
# ─────────────────────────────────────────
COLOR = {
    'header_dark':   '1F3864',
    'header_mid':    '2F5496',
    'header_light':  'BDD7EE',
    'owned':         'E2EFDA',
    'missing':       'FCE4D6',
    'link_es':       'DDEBF7',
    'white':         'FFFFFF',
    'gray':          'F2F2F2',
    'text_white':    'FFFFFF',
    'text_dark':     '1F1F1F',
    'text_red':      'C00000',
    'text_green':    '375623',
    'warn_yellow':   'FFF2CC',
    'warn_text':     '7F6000',
    'no_ref':        'DDEBF7',
    'no_ref_miss':   'FCE4D6',
    'empty_gray':    'F2F2F2',
}

def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

def font(bold=False, color='1F1F1F', size=10, name='Arial'):
    return Font(bold=bold, color=color, size=size, name=name)

def border_thin():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=False)

def left(wrap=False):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap)

# ─────────────────────────────────────────
# CSV 読み込み（UTF-8 BOM / cp932 / 混合エンコード対応）
# ─────────────────────────────────────────
def read_csv(path):
    raw = open(path, 'rb').read()
    if raw[:3] == b'\xef\xbb\xbf':
        raw = raw[3:]
    try:
        text = raw.decode('utf-8')
    except UnicodeDecodeError:
        lines = raw.splitlines()
        try:
            header = lines[0].decode('utf-8')
        except UnicodeDecodeError:
            header = lines[0].decode('cp932', errors='replace')
        data_lines = [line.decode('cp932', errors='replace') for line in lines[1:] if line]
        text = header + '\n' + '\n'.join(data_lines)

    rows = []
    reader = csv.DictReader(io.StringIO(text))
    for row in reader:
        rows.append(row)
    return rows

# ─────────────────────────────────────────
# ユーティリティ
# ─────────────────────────────────────────
SLOT_COLS = ['Main', 'Sub', 'Range', 'Ammo', 'Head', 'Body', 'Hands', 'Legs', 'Feet',
             'Neck', 'Waist', 'Earring1', 'Earring2', 'Ring1', 'Ring2', 'Back']

def normalize_set_no(s):
    """`'002'` → `2`。変換失敗時は `None`"""
    try:
        return int(str(s).strip())
    except (ValueError, TypeError):
        return None

def get_equipset_refs(macro_row):
    """マクロの全行を走査して `/equipset 番号` を全件収集（重複なし int リスト）"""
    refs = []
    for i in range(1, 7):
        line = macro_row.get(f'行{i}', '') or ''
        if '/equipset' not in line.lower():
            continue
        parts = line.strip().split()
        for idx, p in enumerate(parts):
            if p.lower() == '/equipset' and idx + 1 < len(parts):
                no = normalize_set_no(parts[idx + 1])
                if no is not None and no not in refs:
                    refs.append(no)
    return refs

def build_inventory_set(inv_rows):
    return {row['アイテム名'].strip().lower() for row in inv_rows if row.get('アイテム名')}

def build_es_map(es_rows):
    """セットNo（int）→ 装備セット行 dict"""
    result = {}
    for es in es_rows:
        no = normalize_set_no(es.get('セットNo', ''))
        if no is not None:
            result[no] = es
    return result

def build_es_macro_map(macro_rows):
    """セットNo（int）→ 参照マクロのラベルリスト"""
    es_macro_map = {}
    for m in macro_rows:
        refs  = get_equipset_refs(m)
        title = m.get('タイトル', '') or '（無題）'
        book  = m.get('ブックNo', '')
        mno   = m.get('マクロNo', '')
        src   = m.get('ファイル', m.get('ページNo', ''))
        label = f"[{src}]B{book}#{mno} {title}" if src else f"B{book}#{mno} {title}"
        for set_no in refs:
            es_macro_map.setdefault(set_no, []).append(label)
    return es_macro_map

# ─────────────────────────────────────────
# ヘッダー行共通
# ─────────────────────────────────────────
def write_header(ws, row_num, columns, bg='1F3864', fg='FFFFFF', height=22):
    for col_idx, text in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=text)
        cell.fill      = fill(bg)
        cell.font      = font(bold=True, color=fg, size=10)
        cell.alignment = center()
        cell.border    = border_thin()
    ws.row_dimensions[row_num].height = height

# ─────────────────────────────────────────
# シート1（先頭）: サマリー
# ─────────────────────────────────────────
def sheet_summary(wb, macro_rows, es_rows, inv_set, es_macro_map):
    ws = wb.create_sheet('サマリー', 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 40

    ws.merge_cells('A1:C1')
    t = ws['A1']
    t.value      = 'FF11 マクロ・装備整理 サマリー'
    t.fill       = fill(COLOR['header_dark'])
    t.font       = font(bold=True, color=COLOR['text_white'], size=14)
    t.alignment  = center()
    ws.row_dimensions[1].height = 32

    def write_row(r, label, val, note='', bg=COLOR['white']):
        for c in [1, 2, 3]:
            ws.cell(row=r, column=c).fill   = fill(bg)
            ws.cell(row=r, column=c).border = border_thin()
        ws.cell(row=r, column=1, value=label).font = font(bold=True, size=10)
        ws.cell(row=r, column=1).alignment  = left()
        ws.cell(row=r, column=2, value=val).font = font(size=10)
        ws.cell(row=r, column=2).alignment  = center()
        ws.cell(row=r, column=3, value=note).font = font(size=9, color='595959')
        ws.cell(row=r, column=3).alignment  = left()
        ws.row_dimensions[r].height = 18

    def section_header(r, text):
        ws.merge_cells(f'A{r}:C{r}')
        cell = ws.cell(row=r, column=1, value=text)
        cell.fill      = fill(COLOR['header_light'])
        cell.font      = font(bold=True, size=10, color=COLOR['header_dark'])
        cell.alignment = left()
        cell.border    = border_thin()
        ws.row_dimensions[r].height = 20

    r = 2
    section_header(r, '■ マクロ統計'); r += 1
    write_row(r, 'マクロ総数', len(macro_rows)); r += 1
    es_macros = [m for m in macro_rows if get_equipset_refs(m)]
    write_row(r, 'うち装備セット呼出', len(es_macros), '/equipset を含むマクロ'); r += 1

    r += 1
    section_header(r, '■ 装備セット ステータス内訳'); r += 1
    write_row(r, '装備セット総数', len(es_rows)); r += 1

    cnt = {'正常': 0, '装備不足': 0, '未参照': 0, '未参照+不足': 0, '空': 0}
    for es in es_rows:
        set_no    = normalize_set_no(es.get('セットNo', ''))
        items     = [es.get(s, '').strip() for s in SLOT_COLS]
        non_empty = [it for it in items if it]
        if not non_empty:
            cnt['空'] += 1
            continue
        missing    = [it for it in non_empty if it.lower() not in inv_set]
        macros     = es_macro_map.get(set_no, []) if set_no is not None else []
        has_macros = bool(macros)
        has_miss   = bool(missing)
        if   has_macros and not has_miss: cnt['正常']       += 1
        elif has_macros and has_miss:     cnt['装備不足']    += 1
        elif not has_macros and not has_miss: cnt['未参照']  += 1
        else:                             cnt['未参照+不足'] += 1

    write_row(r, '✓ 正常',          cnt['正常'],         '全所持・マクロあり',         bg=COLOR['owned']);       r += 1
    write_row(r, '⚠ 装備不足',      cnt['装備不足'],     'マクロあり・未所持装備あり', bg=COLOR['warn_yellow']); r += 1
    write_row(r, '📦 未参照',        cnt['未参照'],       '装備揃い・マクロから未参照', bg=COLOR['no_ref']);      r += 1
    write_row(r, '🗑 未参照+不足',   cnt['未参照+不足'],  '整理/削除候補',              bg=COLOR['missing']);     r += 1
    write_row(r, '（空）',           cnt['空'],           '空の装備セット');             r += 1

    r += 1
    section_header(r, '■ 所持アイテム'); r += 1
    write_row(r, '総所持アイテム数', len(inv_set), '重複なし・バッグ合計'); r += 1

# ─────────────────────────────────────────
# シート2: 整理チェックシート（新規）
# ─────────────────────────────────────────
_STATUS_ORDER = {'🗑 未参照+不足': 0, '📦 未参照': 1, '⚠ 装備不足': 2, '✓ 正常': 3, '（空）': 4}
_STATUS_STYLE = {
    '✓ 正常':         (COLOR['owned'],      COLOR['text_green']),
    '⚠ 装備不足':     (COLOR['warn_yellow'], COLOR['warn_text']),
    '📦 未参照':      (COLOR['no_ref'],      COLOR['header_dark']),
    '🗑 未参照+不足':  (COLOR['no_ref_miss'], COLOR['text_red']),
    '（空）':          (COLOR['empty_gray'],  '595959'),
}

def sheet_cleanup(wb, es_rows, inv_set, es_macro_map, sheet_name='整理チェック'):
    ws = wb.create_sheet(sheet_name)

    ws.merge_cells('A1:G1')
    t = ws['A1']
    t.value     = '装備セット 整理チェックシート'
    t.fill      = fill(COLOR['header_dark'])
    t.font      = font(bold=True, color=COLOR['text_white'], size=13)
    t.alignment = center()
    ws.row_dimensions[1].height = 28

    cols = ['ステータス', 'セットNo', 'セット名', '参照マクロ数', '参照マクロ一覧', '未所持数', '未所持装備']
    write_header(ws, 2, cols, bg=COLOR['header_mid'])
    for i, w in enumerate([20, 8, 22, 12, 50, 10, 42], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    rows_data = []
    for es in es_rows:
        set_no    = normalize_set_no(es.get('セットNo', ''))
        set_name  = es.get('セット名', '')
        items     = [es.get(s, '').strip() for s in SLOT_COLS]
        non_empty = [it for it in items if it]
        macros    = es_macro_map.get(set_no, []) if set_no is not None else []

        if not non_empty:
            status, missing = '（空）', []
        else:
            missing    = [it for it in non_empty if it.lower() not in inv_set]
            has_macros = bool(macros)
            has_miss   = bool(missing)
            if   has_macros and not has_miss:     status = '✓ 正常'
            elif has_macros and has_miss:          status = '⚠ 装備不足'
            elif not has_macros and not has_miss:  status = '📦 未参照'
            else:                                  status = '🗑 未参照+不足'

        rows_data.append({
            'status':   status,
            'sort_key': _STATUS_ORDER.get(status, 99),
            'set_no':   set_no if set_no is not None else 9999,
            'set_name': set_name,
            'macros':   macros,
            'missing':  missing,
        })

    rows_data.sort(key=lambda x: (x['sort_key'], x['set_no']))

    row_num = 3
    for r in rows_data:
        status = r['status']
        bg_c, fg_c = _STATUS_STYLE.get(status, (COLOR['white'], COLOR['text_dark']))
        macro_text  = '\n'.join(r['macros']) if r['macros'] else '—'
        miss_text   = '、'.join(r['missing']) if r['missing'] else '—'
        line_count  = macro_text.count('\n') + 1

        values = [
            status,
            r['set_no'] if r['set_no'] != 9999 else '',
            r['set_name'],
            len(r['macros']),
            macro_text,
            len(r['missing']),
            miss_text,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill   = fill(bg_c)
            cell.font   = font(size=9, color=fg_c, bold=(col_idx == 1))
            cell.border = border_thin()
            if col_idx in (5, 7):
                cell.alignment = left(wrap=True)
            elif col_idx in (1, 4, 6):
                cell.alignment = center()
            else:
                cell.alignment = left()

        ws.row_dimensions[row_num].height = max(16, min(80, 16 * line_count))
        row_num += 1

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(len(cols))}{row_num - 1}'

# ─────────────────────────────────────────
# シート3: マクロ一覧
# ─────────────────────────────────────────
def sheet_macros(wb, macro_rows, es_map, inv_set):
    ws = wb.create_sheet('マクロ一覧')

    ws.merge_cells('A1:M1')
    title = ws['A1']
    title.value     = 'FF11 マクロ一覧'
    title.fill      = fill(COLOR['header_dark'])
    title.font      = font(bold=True, color=COLOR['text_white'], size=13)
    title.alignment = center()
    ws.row_dimensions[1].height = 28

    cols = ['ページ', '種別', 'マクロNo', 'タイトル',
            '行1', '行2', '行3', '行4', '行5', '行6',
            '参照セットNo', '参照セット名', '装備所持確認']
    write_header(ws, 2, cols, bg=COLOR['header_mid'])
    for i, w in enumerate([12, 8, 8, 22, 30, 30, 30, 20, 20, 20, 14, 30, 24], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _BOOK_LABEL = {'1': '通常', '2': 'Ctrl'}

    row_num = 3
    for m in macro_rows:
        body_lines = [m.get(f'行{i}', '') for i in range(1, 7)]
        es_refs    = get_equipset_refs(m)

        es_nos      = ', '.join(str(n) for n in es_refs) if es_refs else ''
        es_names    = [es_map.get(no, {}).get('セット名', f'セット{no}') for no in es_refs]
        es_name_str = ', '.join(es_names)

        all_missing = []
        for no in es_refs:
            es_info = es_map.get(no, {})
            missing = [es_info.get(s, '') for s in SLOT_COLS
                       if es_info.get(s, '') and es_info.get(s, '').lower() not in inv_set]
            all_missing.extend(missing)

        if not es_refs:
            es_status = ''
        elif not all_missing:
            es_status = '✓ 全所持'
        else:
            uniq = list(dict.fromkeys(all_missing))
            es_status = f'✗ 未所持: {", ".join(uniq[:3])}{"..." if len(uniq) > 3 else ""}'

        row_bg   = COLOR['gray'] if row_num % 2 == 0 else COLOR['white']
        src      = m.get('ファイル', m.get('ページNo', ''))
        book_raw = str(m.get('ブックNo', ''))
        book_lbl = _BOOK_LABEL.get(book_raw, book_raw)
        values = [src, book_lbl, m.get('マクロNo', ''), m.get('タイトル', '')] \
               + body_lines + [es_nos, es_name_str, es_status]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font      = font(size=9)
            cell.alignment = left()
            cell.border    = border_thin()

            if col_idx in range(5, 11) and '/equipset' in str(val).lower():
                cell.fill = fill(COLOR['link_es'])
            elif col_idx == 12 and es_name_str:
                cell.fill = fill(COLOR['link_es'])
                cell.font = font(size=9, bold=True, color='1F3864')
            elif col_idx == 13 and es_status.startswith('✓'):
                cell.fill = fill(COLOR['owned'])
                cell.font = font(size=9, color=COLOR['text_green'])
            elif col_idx == 13 and es_status.startswith('✗'):
                cell.fill = fill(COLOR['missing'])
                cell.font = font(size=9, color=COLOR['text_red'])
            else:
                cell.fill = fill(row_bg)

        ws.row_dimensions[row_num].height = 16
        row_num += 1

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(len(cols))}{row_num - 1}'

# ─────────────────────────────────────────
# シート4: 装備セット × 所持確認
# ─────────────────────────────────────────
def sheet_equipsets(wb, es_rows, inv_set):
    ws = wb.create_sheet('装備セット確認')

    ws.merge_cells(f'A1:{get_column_letter(2 + len(SLOT_COLS))}1')
    title = ws['A1']
    title.value     = 'FF11 装備セット × 所持確認'
    title.fill      = fill(COLOR['header_dark'])
    title.font      = font(bold=True, color=COLOR['text_white'], size=13)
    title.alignment = center()
    ws.row_dimensions[1].height = 28

    cols = ['セットNo', 'セット名'] + SLOT_COLS + ['未所持数', '未所持アイテム']
    write_header(ws, 2, cols, bg=COLOR['header_mid'])
    for i, w in enumerate([8, 20] + [18] * 16 + [10, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row_num = 3
    for es in es_rows:
        row_bg        = COLOR['gray'] if row_num % 2 == 0 else COLOR['white']
        missing_items = []
        item_cells    = []

        for slot in SLOT_COLS:
            item  = es.get(slot, '').strip()
            owned = (not item) or (item.lower() in inv_set)
            item_cells.append((item, owned))
            if item and not owned:
                missing_items.append(item)

        values = [es.get('セットNo', ''), es.get('セット名', '')]
        values += [ic[0] for ic in item_cells]
        values += [len(missing_items), ', '.join(missing_items)]

        sc_s = 3
        sc_e = sc_s + len(SLOT_COLS) - 1
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font      = font(size=9)
            cell.alignment = left()
            cell.border    = border_thin()

            if sc_s <= col_idx <= sc_e:
                item, owned = item_cells[col_idx - sc_s]
                if not item:
                    cell.fill = fill(row_bg)
                elif owned:
                    cell.fill = fill(COLOR['owned'])
                    cell.font = font(size=9, color=COLOR['text_green'])
                else:
                    cell.fill = fill(COLOR['missing'])
                    cell.font = font(size=9, color=COLOR['text_red'], bold=True)
            elif col_idx == sc_e + 1:
                cell.fill      = fill(COLOR['owned'] if not missing_items else COLOR['missing'])
                cell.font      = font(size=9, bold=True,
                                      color=COLOR['text_green'] if not missing_items else COLOR['text_red'])
                cell.alignment = center()
            elif col_idx == sc_e + 2:
                cell.fill      = fill(row_bg)
                cell.alignment = left(wrap=True)
            else:
                cell.fill = fill(row_bg)

        ws.row_dimensions[row_num].height = 16
        row_num += 1

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(len(cols))}{row_num - 1}'

# ─────────────────────────────────────────
# シート5: 所持アイテム一覧
# ─────────────────────────────────────────
def sheet_inventory(wb, inv_rows, sheet_name='所持アイテム一覧'):
    ws = wb.create_sheet(sheet_name)

    ws.merge_cells('A1:F1')
    title = ws['A1']
    title.value     = 'FF11 所持アイテム一覧'
    title.fill      = fill(COLOR['header_dark'])
    title.font      = font(bold=True, color=COLOR['text_white'], size=13)
    title.alignment = center()
    ws.row_dimensions[1].height = 28

    cols = ['バッグ', 'スロット', 'アイテムID', 'アイテム名', '個数', 'オーグメント']
    write_header(ws, 2, cols, bg=COLOR['header_mid'])
    for i, w in enumerate([14, 8, 10, 30, 8, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    bag_colors = {
        'Inventory': 'DEEAF1', 'Safe': 'E2EFDA', 'Storage': 'FFF2CC',
        'Wardrobe1': 'FCE4D6', 'Wardrobe2': 'F4B8B8', 'Wardrobe3': 'E8D5F5',
        'Wardrobe4': 'D5E8D4', 'Wardrobe5': 'DAE8FC',
    }

    row_num = 3
    for item in inv_rows:
        bag    = item.get('バッグ', '')
        bg     = bag_colors.get(bag, COLOR['gray'])
        values = [bag, item.get('スロット', ''), item.get('アイテムID', ''),
                  item.get('アイテム名', ''), item.get('個数', ''), item.get('オーグメント', '')]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill      = fill(bg)
            cell.font      = font(size=9)
            cell.alignment = left()
            cell.border    = border_thin()
        ws.row_dimensions[row_num].height = 15
        row_num += 1

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:F{row_num - 1}'

# ─────────────────────────────────────────
# メイン
# ─────────────────────────────────────────
def find_latest_csv(pattern):
    files = glob.glob(pattern)
    return sorted(files)[-1] if files else None

def find_latest_csv_per_char(export_dir, suffix):
    """キャラ名ごとに最新の {char}_{suffix}_YYYYMMDD_HHMMSS.csv を返す辞書"""
    pattern = os.path.join(export_dir, f'*_{suffix}_*.csv')
    latest  = {}
    for path in sorted(glob.glob(pattern)):
        m = re.match(rf'^(.+?)_{re.escape(suffix)}_\d{{8}}_\d{{6}}\.csv$',
                     os.path.basename(path))
        if not m:
            continue
        char = m.group(1)
        if char not in latest or \
           os.path.getmtime(path) > os.path.getmtime(latest[char]):
            latest[char] = path
    return latest

BAG_ORDER = ['Inventory', 'Safe', 'Storage', 'Satchel', 'Sack', 'Case',
             'Wardrobe1', 'Wardrobe2', 'Wardrobe3', 'Wardrobe4', 'Wardrobe5',
             'Wardrobe6', 'Wardrobe7', 'Wardrobe8',
             'Locker', 'Safe2', 'Temporary']

def sheet_all_chars_summary(wb, char_data_list):
    """全キャラのインベントリ概要シートを Workbook 先頭に挿入"""
    ws = wb.create_sheet('全キャラサマリー', 0)
    ws.sheet_view.showGridLines = False

    cols = ['キャラ名', '合計'] + BAG_ORDER
    total_cols = len(cols)

    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    t = ws['A1']
    t.value     = 'FF11 全キャラ 所持アイテムサマリー'
    t.fill      = fill(COLOR['header_dark'])
    t.font      = font(bold=True, color=COLOR['text_white'], size=14)
    t.alignment = center()
    ws.row_dimensions[1].height = 32

    write_header(ws, 2, cols, bg=COLOR['header_mid'])

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 8
    for i in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 10

    row_num = 3
    totals  = {b: 0 for b in BAG_ORDER}
    for entry in sorted(char_data_list, key=lambda x: x['char']):
        bg     = COLOR['white'] if row_num % 2 == 1 else COLOR['gray']
        bags   = entry['bag_counts']
        values = [entry['char'], entry['inv_count']] + [bags.get(b, 0) for b in BAG_ORDER]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill      = fill(bg)
            cell.font      = font(size=10, bold=(col_idx == 1))
            cell.alignment = left() if col_idx == 1 else center()
            cell.border    = border_thin()
        ws.row_dimensions[row_num].height = 18
        for b in BAG_ORDER:
            totals[b] += bags.get(b, 0)
        row_num += 1

    # 合計行
    total_values = ['合計', sum(entry['inv_count'] for entry in char_data_list)] \
                 + [totals[b] for b in BAG_ORDER]
    for col_idx, val in enumerate(total_values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.fill      = fill(COLOR['header_light'])
        cell.font      = font(size=10, bold=True, color=COLOR['header_dark'])
        cell.alignment = left() if col_idx == 1 else center()
        cell.border    = border_thin()
    ws.row_dimensions[row_num].height = 20

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(total_cols)}{row_num - 1}'

def run_single():
    """単一キャラ処理（既存の main() 相当、後方互換用）"""
    base = os.path.dirname(os.path.abspath(__file__))

    if len(sys.argv) >= 4 and not sys.argv[1].startswith('--'):
        macro_path = sys.argv[1]
        es_path    = sys.argv[2]
        inv_path   = sys.argv[3]
    else:
        char_prefix = ''
        if '--char' in sys.argv:
            idx = sys.argv.index('--char')
            if idx + 1 < len(sys.argv):
                char_prefix = sys.argv[idx + 1] + '_'

        macro_path = find_latest_csv(os.path.join(base, f'{char_prefix}*macros_*.csv'))
        es_path    = find_latest_csv(os.path.join(base, f'{char_prefix}*equipsets_*.csv'))
        inv_path   = find_latest_csv(os.path.join(base, f'{char_prefix}*inventory_*.csv'))

        missing = []
        if not macro_path: missing.append(f'{char_prefix}*macros_*.csv')
        if not es_path:    missing.append(f'{char_prefix}*equipsets_*.csv')
        if not inv_path:   missing.append(f'{char_prefix}*inventory_*.csv')
        if missing:
            print(f'[エラー] 以下のファイルが見つかりません: {", ".join(missing)}')
            sys.exit(1)

        print('[使用ファイル]')
        print(f'  マクロ      : {macro_path}')
        print(f'  装備セット  : {es_path}')
        print(f'  所持アイテム: {inv_path}')

    macro_rows   = read_csv(macro_path)
    es_rows      = read_csv(es_path)
    inv_rows     = read_csv(inv_path)
    inv_set      = build_inventory_set(inv_rows)
    es_map       = build_es_map(es_rows)
    es_macro_map = build_es_macro_map(macro_rows)

    wb = Workbook()
    wb.remove(wb.active)

    sheet_summary(wb, macro_rows, es_rows, inv_set, es_macro_map)
    sheet_cleanup(wb, es_rows, inv_set, es_macro_map)
    sheet_macros(wb, macro_rows, es_map, inv_set)
    sheet_equipsets(wb, es_rows, inv_set)
    sheet_inventory(wb, inv_rows)

    out_path = os.path.join(base, 'FF11_macro_review.xlsx')
    wb.save(out_path)
    print(f'\nOK Excel出力完了: {out_path}')
    return out_path

def run_multi(export_dir):
    """複数キャラ処理: export_dir 内の最新 CSV を全キャラ分まとめて Excel 化"""
    inv_map   = find_latest_csv_per_char(export_dir, 'inventory')
    macro_map = find_latest_csv_per_char(export_dir, 'macros')
    es_map_f  = find_latest_csv_per_char(export_dir, 'equipsets')

    if not inv_map:
        print(f'[エラー] {export_dir} に *_inventory_*.csv が見つかりません')
        sys.exit(1)

    print(f'[検出キャラ] {sorted(inv_map.keys())}')

    wb = Workbook()
    wb.remove(wb.active)

    char_data_list = []
    used_names     = set()

    for char_name in sorted(inv_map.keys()):
        inv_path = inv_map[char_name]
        print(f'\n--- {char_name} ---')
        print(f'  inventory : {inv_path}')

        inv_rows   = read_csv(inv_path)
        bag_counts = {}
        for row in inv_rows:
            bag = row.get('バッグ', '')
            bag_counts[bag] = bag_counts.get(bag, 0) + 1

        char_data_list.append({
            'char':       char_name,
            'inv_count':  len(inv_rows),
            'bag_counts': bag_counts,
        })

        # Excel シート名の衝突回避（禁止文字除去・31文字制限）
        safe = re.sub(r'[\\/:*?"<>|\[\]]', '_', char_name)
        base_name = safe[:27] + '_持物'
        sheet_nm  = base_name
        suffix_n  = 2
        while sheet_nm in used_names:
            sheet_nm = safe[:25] + f'_{suffix_n}_持物'
            suffix_n += 1
        used_names.add(sheet_nm)
        sheet_inventory(wb, inv_rows, sheet_name=sheet_nm)

        # マクロ・装備セット CSV が揃っていれば整理シートも追加
        mp = macro_map.get(char_name)
        ep = es_map_f.get(char_name)
        if mp and ep:
            print(f'  macros    : {mp}')
            print(f'  equipsets : {ep}')
            macro_rows   = read_csv(mp)
            es_rows      = read_csv(ep)
            inv_set      = build_inventory_set(inv_rows)
            es_mac_map   = build_es_macro_map(macro_rows)
            cln_name     = re.sub(r'[\\/:*?"<>|\[\]]', '_', char_name)[:27] + '_整理'
            suffix_n2    = 2
            while cln_name in used_names:
                cln_name = re.sub(r'[\\/:*?"<>|\[\]]', '_', char_name)[:24] + f'_{suffix_n2}_整理'
                suffix_n2 += 1
            used_names.add(cln_name)
            sheet_cleanup(wb, es_rows, inv_set, es_mac_map, sheet_name=cln_name)

    sheet_all_chars_summary(wb, char_data_list)

    out_path = os.path.join(export_dir, 'FF11_all_chars.xlsx')
    wb.save(out_path)
    print(f'\nOK Excel出力完了: {out_path}')

    gdrive_dir = r"G:\マイドライブ\Ashitav4\export"
    if os.path.isdir(gdrive_dir):
        gdrive_path = os.path.join(gdrive_dir, 'FF11_all_chars.xlsx')
        shutil.copy2(out_path, gdrive_path)
        print(f'OK Google Drive コピー完了: {gdrive_path}')
    else:
        print(f'INFO Google Drive フォルダが見つからないためスキップ: {gdrive_dir}')

    return out_path

def main():
    if '--multi' in sys.argv:
        export_dir = os.path.dirname(os.path.abspath(__file__))
        if '--dir' in sys.argv:
            idx = sys.argv.index('--dir')
            if idx + 1 < len(sys.argv):
                export_dir = sys.argv[idx + 1]
        run_multi(export_dir)
    else:
        run_single()

if __name__ == '__main__':
    main()
