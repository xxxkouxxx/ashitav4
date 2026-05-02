"""
pupsets_export.py  - からくり師アタッチメントセット Excel 出力スクリプト
使い方:
  python pupsets_export.py              # 保存済みセット（config/addons/pupsets/）
  python pupsets_export.py all          # アドオン内全セット（addons/pupsets/sets/）
  python pupsets_export.py [フォルダ] [出力xlsx]  # フォルダ直接指定

引数省略時はデフォルトパスを使用:
  setsフォルダ : C:\\Ashita-v4beta\\config\\addons\\pupsets
  出力xlsx     : pupsets_export.xlsx（スクリプトと同じフォルダ）
"""

import sys
import os
import glob
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────
# 定数
# ─────────────────────────────────────────
DEFAULT_SETS_DIR  = r'C:\Ashita-v4beta\config\addons\pupsets'
ALL_SETS_DIR      = r'C:\Ashita-v4beta\addons\pupsets\sets'
# 読み込み対象外ファイル（テンプレート等）
SKIP_FILES = {'TEMPLATE_DO_NOT_LOAD'}

SLOT_LABELS = [
    'Head', 'Frame',
    'Att  1', 'Att  2', 'Att  3', 'Att  4',
    'Att  5', 'Att  6', 'Att  7', 'Att  8',
    'Att  9', 'Att 10', 'Att 11', 'Att 12',
]

COLOR = {
    'header':     '1F3864',  # 濃紺（ヘッダ行）
    'head_frame': 'BDD7EE',  # 薄青（Head / Frame 行）
    'att':        'FFFFFF',  # 白（アタッチメント奇数行）
    'att_alt':    'F2F2F2',  # 薄グレー（アタッチメント偶数行）
    'empty':      'FCE4D6',  # 薄赤（未装着セル）
    'text_white': 'FFFFFF',
    'text_dark':  '1F1F1F',
}

# ─────────────────────────────────────────
# スタイルヘルパー
# ─────────────────────────────────────────
def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

def font(bold=False, color='1F1F1F', size=10, name='Yu Gothic'):
    return Font(bold=bold, color=color, size=size, name=name)

def border_thin():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def align_center():
    return Alignment(horizontal='center', vertical='center')

def align_left():
    return Alignment(horizontal='left', vertical='center')

def style_cell(cell, bg, bold=False, text_color='1F1F1F', align='left'):
    cell.fill = fill(bg)
    cell.font = font(bold=bold, color=text_color)
    cell.border = border_thin()
    cell.alignment = align_center() if align == 'center' else align_left()

# ─────────────────────────────────────────
# セットファイル読み込み
# ─────────────────────────────────────────
def read_set(path: str) -> list:
    """cp932 で .txt を読み、14要素リストを返す（空行 / 不足行は '---'）
    JSON 形式（addons/pupsets/sets/）と1行1アイテム形式（config/addons/pupsets/）の両方に対応"""
    try:
        raw = open(path, 'rb').read()
        text = raw.decode('cp932', errors='replace')
    except OSError:
        return ['---'] * 14

    # JSON 形式の判定（先頭が { ）
    if text.lstrip().startswith('{'):
        try:
            # cp932 でデコード済みのテキストを再度 JSON パース
            data = json.loads(text)
            head  = data.get('head', '---') or '---'
            frame = data.get('frame', '---') or '---'
            atts  = data.get('attachments', [])
            result = [head, frame]
            for a in atts[:12]:
                result.append(a if a and a.strip() else '---')
            while len(result) < 14:
                result.append('---')
            return result
        except json.JSONDecodeError:
            return ['---'] * 14

    # 1行1アイテム形式
    lines = text.splitlines()
    result = []
    for ln in lines[:14]:
        result.append(ln.strip() if ln.strip() else '---')
    while len(result) < 14:
        result.append('---')
    return result

def load_all_sets(folder: str) -> dict:
    """フォルダ内 *.txt を全読み込み → {セット名: [14行]} をアルファベット順で返す
    SKIP_FILES に含まれるファイル名はスキップ"""
    pattern = os.path.join(folder, '*.txt')
    paths = sorted(glob.glob(pattern))
    sets = {}
    for p in paths:
        name = os.path.splitext(os.path.basename(p))[0]
        if name in SKIP_FILES:
            continue
        sets[name] = read_set(p)
    return sets

# ─────────────────────────────────────────
# Excel シート生成
# ─────────────────────────────────────────
def build_sheet(wb: Workbook, sets: dict):
    ws = wb.active
    ws.title = 'Pupsets'

    set_names = list(sets.keys())
    total_cols = 1 + len(set_names)  # Slot列 + セット数

    # ── ヘッダ行 ──────────────────────────
    ws.cell(1, 1, 'Slot')
    style_cell(ws.cell(1, 1), COLOR['header'], bold=True,
               text_color=COLOR['text_white'], align='center')

    for col_idx, sname in enumerate(set_names, start=2):
        c = ws.cell(1, col_idx, sname)
        style_cell(c, COLOR['header'], bold=True,
                   text_color=COLOR['text_white'], align='center')

    # ── データ行 ──────────────────────────
    for row_idx, label in enumerate(SLOT_LABELS, start=2):
        slot_num = row_idx - 2  # 0=Head, 1=Frame, 2〜13=Att

        # スロットラベル列の背景色決定
        if slot_num < 2:
            row_bg = COLOR['head_frame']
        elif slot_num % 2 == 0:
            row_bg = COLOR['att']
        else:
            row_bg = COLOR['att_alt']

        # Slot ラベルセル
        lc = ws.cell(row_idx, 1, label)
        style_cell(lc, row_bg, bold=True, align='center')

        # 各セットの値
        for col_idx, sname in enumerate(set_names, start=2):
            value = sets[sname][slot_num]
            cell_bg = COLOR['empty'] if value == '---' else row_bg
            c = ws.cell(row_idx, col_idx, value)
            style_cell(c, cell_bg)

    # ── 列幅調整 ──────────────────────────
    ws.column_dimensions['A'].width = 10  # Slot 列
    for col_idx in range(2, total_cols + 1):
        # セット名と全セル値から最大文字数を推定（日本語は1.8倍換算）
        sname = set_names[col_idx - 2]
        max_len = len(sname)
        for row_idx in range(2, 16):
            val = ws.cell(row_idx, col_idx).value or ''
            # 日本語文字はおよそ2文字分の幅
            char_w = sum(2 if ord(ch) > 127 else 1 for ch in val)
            max_len = max(max_len, char_w)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    # ── 行の高さ ──────────────────────────
    ws.row_dimensions[1].height = 18
    for row_idx in range(2, 16):
        ws.row_dimensions[row_idx].height = 16

    # ウィンドウ枠の固定（ヘッダ行 + Slot 列）
    ws.freeze_panes = 'B2'

# ─────────────────────────────────────────
# エントリポイント
# ─────────────────────────────────────────
def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    arg1 = sys.argv[1] if len(sys.argv) >= 2 else ''

    # "all" 指定時は addons/pupsets/sets/ を対象にする
    if arg1.lower() == 'all':
        sets_dir = ALL_SETS_DIR
        default_out = os.path.join(script_dir, 'pupsets_all.xlsx')
        out_path = sys.argv[2] if len(sys.argv) >= 3 else default_out
    else:
        sets_dir = arg1 if arg1 else DEFAULT_SETS_DIR
        out_path = sys.argv[2] if len(sys.argv) >= 3 else os.path.join(script_dir, 'pupsets_export.xlsx')

    print(f'読み込みフォルダ : {sets_dir}')
    sets = load_all_sets(sets_dir)
    if not sets:
        print('エラー: セットファイルが見つかりません')
        sys.exit(1)

    print(f'セット数         : {len(sets)} ({", ".join(sets.keys())})')

    wb = Workbook()
    build_sheet(wb, sets)
    wb.save(out_path)
    print(f'出力完了         : {out_path}')

if __name__ == '__main__':
    main()
